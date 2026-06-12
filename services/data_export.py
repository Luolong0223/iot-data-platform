"""
数据导出服务 - 支持 Excel、CSV 格式导出
"""
import io
import csv
import logging
from datetime import datetime, timedelta
from flask import Response, make_response

from models.database import db, Device, SlaveChannel, DataPoint, AlarmRecord, User

logger = logging.getLogger(__name__)


class DataExportService:
    """数据导出服务类"""
    
    @staticmethod
    def export_data_points_to_csv(user_id, device_id=None, channel_id=None, 
                                    start_time=None, end_time=None, limit=10000):
        """
        导出数据点到CSV
        
        Args:
            user_id: 用户ID
            device_id: 设备ID（可选）
            channel_id: 通道ID（可选）
            start_time: 开始时间（可选）
            end_time: 结束时间（可选）
            limit: 最大导出数量
        """
        # 构建查询
        query = DataPoint.query.join(SlaveChannel).join(Device).filter(
            Device.user_id == user_id
        )
        
        if device_id:
            query = query.filter(Device.id == device_id)
        if channel_id:
            query = query.filter(SlaveChannel.id == channel_id)
        if start_time:
            query = query.filter(DataPoint.timestamp >= start_time)
        if end_time:
            query = query.filter(DataPoint.timestamp <= end_time)
        
        query = query.order_by(DataPoint.timestamp.desc()).limit(limit)
        data_points = query.all()
        
        # 创建CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['ID', '设备名称', '通道名称', '数据点名称', '数值', '时间'])
        
        # 写入数据
        for dp in data_points:
            writer.writerow([
                dp.id,
                dp.channel.device.name if dp.channel and dp.channel.device else '',
                dp.channel.name if dp.channel else '',
                dp.name,
                dp.value,
                dp.timestamp.strftime('%Y-%m-%d %H:%M:%S') if dp.timestamp else ''
            ])
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_data_points_to_excel(user_id, device_id=None, channel_id=None,
                                      start_time=None, end_time=None, limit=10000):
        """
        导出数据点到Excel
        
        Returns:
            bytes: Excel文件内容
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        # 构建查询
        query = DataPoint.query.join(SlaveChannel).join(Device).filter(
            Device.user_id == user_id
        )
        
        if device_id:
            query = query.filter(Device.id == device_id)
        if channel_id:
            query = query.filter(SlaveChannel.id == channel_id)
        if start_time:
            query = query.filter(DataPoint.timestamp >= start_time)
        if end_time:
            query = query.filter(DataPoint.timestamp <= end_time)
        
        query = query.order_by(DataPoint.timestamp.desc()).limit(limit)
        data_points = query.all()
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '数据导出'
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        headers = ['ID', '设备名称', '通道名称', '数据点名称', '数值', '时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row, dp in enumerate(data_points, 2):
            ws.cell(row=row, column=1, value=dp.id)
            ws.cell(row=row, column=2, value=dp.channel.device.name if dp.channel and dp.channel.device else '')
            ws.cell(row=row, column=3, value=dp.channel.name if dp.channel else '')
            ws.cell(row=row, column=4, value=dp.name)
            ws.cell(row=row, column=5, value=dp.value)
            ws.cell(row=row, column=6, value=dp.timestamp.strftime('%Y-%m-%d %H:%M:%S') if dp.timestamp else '')
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_devices_to_excel(user_id):
        """导出设备列表到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        devices = Device.query.filter_by(user_id=user_id).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '设备列表'
        
        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        
        headers = ['ID', '设备名称', '设备类型', '电压(mV)', '纬度', '经度', '位置', '在线状态', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, device in enumerate(devices, 2):
            ws.cell(row=row, column=1, value=device.id)
            ws.cell(row=row, column=2, value=device.name)
            ws.cell(row=row, column=3, value=device.device_type or '')
            ws.cell(row=row, column=4, value=device.voltage_mv)
            ws.cell(row=row, column=5, value=device.latitude)
            ws.cell(row=row, column=6, value=device.longitude)
            ws.cell(row=row, column=7, value=device.location_name or '')
            ws.cell(row=row, column=8, value='在线' if device.is_online else '离线')
            ws.cell(row=row, column=9, value=device.created_at.strftime('%Y-%m-%d %H:%M:%S') if device.created_at else '')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_alarms_to_excel(user_id, start_time=None, end_time=None, is_read=None):
        """导出报警记录到Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        
        query = AlarmRecord.query.filter_by(user_id=user_id)
        
        if start_time:
            query = query.filter(AlarmRecord.created_at >= start_time)
        if end_time:
            query = query.filter(AlarmRecord.created_at <= end_time)
        if is_read is not None:
            query = query.filter_by(is_read=is_read)
        
        alarms = query.order_by(AlarmRecord.created_at.desc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '报警记录'
        
        # 表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
        
        headers = ['ID', '设备名称', '通道名称', '数据点', '数值', '阈值', '条件', '严重程度', '消息', '已读', '创建时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, alarm in enumerate(alarms, 2):
            ws.cell(row=row, column=1, value=alarm.id)
            ws.cell(row=row, column=2, value=alarm.device_name)
            ws.cell(row=row, column=3, value=alarm.channel_name)
            ws.cell(row=row, column=4, value=alarm.point_name)
            ws.cell(row=row, column=5, value=alarm.value)
            ws.cell(row=row, column=6, value=alarm.threshold)
            ws.cell(row=row, column=7, value=alarm.condition)
            ws.cell(row=row, column=8, value=alarm.severity)
            ws.cell(row=row, column=9, value=alarm.message)
            ws.cell(row=row, column=10, value='是' if alarm.is_read else '否')
            ws.cell(row=row, column=11, value=alarm.created_at.strftime('%Y-%m-%d %H:%M:%S') if alarm.created_at else '')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def generate_response(content, filename, content_type):
        """生成下载响应"""
        response = make_response(content)
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
